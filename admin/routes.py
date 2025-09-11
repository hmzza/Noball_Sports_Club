"""
Main admin routes module - consolidated and clean.
"""
from flask import Blueprint, render_template, request, jsonify, session, redirect, url_for, g
import logging

from auth_middleware import require_auth, require_permission, super_admin_only, admin_or_higher, check_current_user
from services.auth_service import AuthService, SessionManager
from services.activity_service import ActivityService
from services.admin_service import AdminService
from .views import AdminDashboardView, AdminBookingView, AdminScheduleView, AdminBookingControlView, AdminPricingView, AdminAPIView, AdminExpenseView
from services.contact_service import ContactService
from datetime import datetime

logger = logging.getLogger(__name__)

admin_bp = Blueprint("admin_panel", __name__, url_prefix="/admin")

# Authentication Routes
@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    """Admin login"""
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        
        user = AuthService.authenticate_user(username, password)
        if user:
            # Store user data directly in Flask session instead of using in-memory SessionManager
            session["admin_logged_in"] = True
            session["admin_user_id"] = user.id
            session["admin_username"] = user.username
            session["admin_role"] = user.role
            return redirect(url_for("admin_panel.admin_dashboard"))
        else:
            return render_template("admin_login.html", error="Invalid credentials")
    
    return render_template("admin_login.html")

@admin_bp.route("/logout")
def logout():
    """Admin logout"""
    # Clear all admin session data
    session.pop("admin_logged_in", None)
    session.pop("admin_user_id", None)
    session.pop("admin_username", None)
    session.pop("admin_role", None)
    return redirect(url_for("admin_panel.login"))

# Check current user before every request
@admin_bp.before_request
def before_request():
    """Check current user before every request"""
    if request.endpoint not in ['admin_panel.login']:
        check_current_user()
        
# Root admin route redirects to login
@admin_bp.route("/")
def admin_root():
    """Admin root - redirect to login or dashboard"""
    if session.get('admin_logged_in'):
        return redirect(url_for('admin_panel.admin_dashboard'))
    return redirect(url_for('admin_panel.login'))

# Unauthorized access page
@admin_bp.route("/unauthorized")
@require_auth
def unauthorized():
    """Show unauthorized access page"""
    return render_template("admin_unauthorized.html")

# Main Admin Views
@admin_bp.route("/dashboard")
@require_auth
@require_permission('view_dashboard')
def admin_dashboard():
    """Admin dashboard"""
    return AdminDashboardView.render_dashboard()

@admin_bp.route("/bookings")
@require_auth
@require_permission('view_bookings')
def admin_bookings():
    """Admin bookings management"""
    return AdminBookingView.render_bookings()

@admin_bp.route("/schedule")
@require_auth
@require_permission('view_schedule')
def admin_schedule():
    """Schedule management"""
    return AdminScheduleView.render_schedule()

@admin_bp.route("/booking-control")
@require_auth
@require_permission('manage_bookings')
def admin_booking_control():
    """Modern booking control center"""
    return AdminBookingControlView.render_booking_control()

@admin_bp.route("/history")
@require_auth
@require_permission('view_bookings')
def admin_history():
    """Comprehensive booking history with filters and search"""
    from .views import AdminHistoryView
    return AdminHistoryView.render_history()

@admin_bp.route("/pricing")
@require_auth
@require_permission('view_pricing')
def admin_pricing():
    """Pricing management"""
    return AdminPricingView.render_pricing()

@admin_bp.route("/promo-codes")
@require_auth
@require_permission('view_promo_codes')
def admin_promo_codes():
    """Promo codes management"""
    return render_template("admin_promo_codes.html")

@admin_bp.route("/expenses")
@require_auth
@require_permission('view_expenses')
def admin_expenses():
    """Expense management"""
    return AdminExpenseView.render_expenses()

@admin_bp.route("/contacts")
@require_auth
@require_permission('view_reports')
def admin_contacts():
    """Admin contacts view"""
    try:
        contacts = ContactService.get_all_contacts()
        return render_template("admin_contacts.html", contacts=contacts)
    except Exception as e:
        logger.error(f"Error loading contacts: {e}")
        return render_template("admin_contacts.html", contacts=[])

# User Management Routes (Super Admin Only)
@admin_bp.route("/users")
@require_auth
@super_admin_only
def admin_users():
    """User management for super admin"""
    try:
        users = AuthService.get_all_users()
        return render_template("admin_users.html", users=users)
    except Exception as e:
        logger.error(f"Error loading users: {e}")
        return render_template("admin_users.html", users=[])

@admin_bp.route("/create-user", methods=["POST"])
@require_auth
@super_admin_only
def create_user():
    """Create new admin user"""
    try:
        username = request.form.get("username")
        password = request.form.get("password")
        role = request.form.get("role")
        
        if not username or not password or not role:
            return jsonify({"error": "All fields are required"}), 400
        
        user = AuthService.create_user(username, password, role, g.current_user.id)
        if user:
            # Log the user creation activity
            ActivityService.log_user_created(str(user.id), username, role)
            return jsonify({"success": True, "message": "User created successfully"})
        else:
            return jsonify({"error": "Username already exists"}), 400
    except Exception as e:
        logger.error(f"Error creating user: {e}")
        return jsonify({"error": "Failed to create user"}), 500

@admin_bp.route("/edit-user/<int:user_id>", methods=["POST"])
@require_auth
@super_admin_only
def edit_user(user_id):
    """Edit user details"""
    try:
        username = request.form.get("username")
        role = request.form.get("role")
        
        updates = {}
        if username:
            updates["username"] = username
        if role:
            updates["role"] = role
        
        if AuthService.update_user(user_id, updates):
            # Log the user update activity
            user = AuthService.get_user_by_id(user_id)
            if user:
                ActivityService.log_activity('user_updated', 'user', str(user_id), user.username, 
                                           f"Updated: {', '.join(updates.keys())}")
            return jsonify({"success": True, "message": "User updated successfully"})
        else:
            return jsonify({"error": "Failed to update user"}), 500
    except Exception as e:
        logger.error(f"Error editing user: {e}")
        return jsonify({"error": "Failed to update user"}), 500

@admin_bp.route("/change-password/<int:user_id>", methods=["POST"])
@require_auth
@super_admin_only
def change_user_password(user_id):
    """Change user password"""
    try:
        new_password = request.form.get("new_password")
        
        if not new_password:
            return jsonify({"error": "New password is required"}), 400
        
        if AuthService.change_password(user_id, new_password):
            # Log the password change activity
            user = AuthService.get_user_by_id(user_id)
            if user:
                ActivityService.log_activity('user_password_changed', 'user', str(user_id), user.username)
            return jsonify({"success": True, "message": "Password changed successfully"})
        else:
            return jsonify({"error": "Failed to change password"}), 500
    except Exception as e:
        logger.error(f"Error changing password: {e}")
        return jsonify({"error": "Failed to change password"}), 500

@admin_bp.route("/lock-user/<int:user_id>", methods=["POST"])
@require_auth
@super_admin_only
def lock_user(user_id):
    """Lock user account"""
    try:
        if AuthService.lock_user(user_id):
            # Log the user lock activity
            user = AuthService.get_user_by_id(user_id)
            if user:
                ActivityService.log_user_locked(str(user_id), user.username)
            return jsonify({"success": True, "message": "User locked successfully"})
        else:
            return jsonify({"error": "Failed to lock user"}), 500
    except Exception as e:
        logger.error(f"Error locking user: {e}")
        return jsonify({"error": "Failed to lock user"}), 500

@admin_bp.route("/unlock-user/<int:user_id>", methods=["POST"])
@require_auth
@super_admin_only
def unlock_user(user_id):
    """Unlock user account"""
    try:
        if AuthService.unlock_user(user_id):
            # Log the user unlock activity
            user = AuthService.get_user_by_id(user_id)
            if user:
                ActivityService.log_user_unlocked(str(user_id), user.username)
            return jsonify({"success": True, "message": "User unlocked successfully"})
        else:
            return jsonify({"error": "Failed to unlock user"}), 500
    except Exception as e:
        logger.error(f"Error unlocking user: {e}")
        return jsonify({"error": "Failed to unlock user"}), 500

# Activity Logs Routes
@admin_bp.route("/logs")
@require_auth
@require_permission('view_logs')
def admin_logs():
    """Activity logs page"""
    try:
        page = int(request.args.get('page', 1))
        limit = 50
        offset = (page - 1) * limit
        
        logs = ActivityService.get_all_logs(limit=limit, offset=offset)
        stats = ActivityService.get_log_stats()
        
        # Get unique usernames for filter
        unique_users = list(set([log.username for log in logs]))
        unique_users.sort()
        
        return render_template("admin_logs.html", 
                             logs=logs, 
                             stats=stats, 
                             unique_users=unique_users,
                             page=page)
    except Exception as e:
        logger.error(f"Error loading logs: {e}")
        return render_template("admin_logs.html", logs=[], stats={}, unique_users=[], page=1)

@admin_bp.route("/delete-log/<int:log_id>", methods=["POST"])
@require_auth
@super_admin_only
def delete_log(log_id):
    """Delete a specific activity log"""
    try:
        if ActivityService.delete_log(log_id):
            return jsonify({"success": True, "message": "Log deleted successfully"})
        else:
            return jsonify({"error": "Failed to delete log"}), 500
    except Exception as e:
        logger.error(f"Error deleting log: {e}")
        return jsonify({"error": "Failed to delete log"}), 500

@admin_bp.route("/cleanup-logs", methods=["POST"])
@require_auth
@super_admin_only
def cleanup_logs():
    """Delete old activity logs"""
    try:
        data = request.get_json()
        days = int(data.get('days', 90))
        
        deleted_count = ActivityService.delete_old_logs(days)
        return jsonify({"success": True, "deleted_count": deleted_count})
    except Exception as e:
        logger.error(f"Error cleaning up logs: {e}")
        return jsonify({"error": "Failed to cleanup logs"}), 500

# Dangerous operations (Super Admin Only)
@admin_bp.route("/cleanup-bookings", methods=["POST"])
@require_auth
@super_admin_only
def cleanup_bookings():
    """Delete old bookings (super admin only). Requires confirmation text."""
    try:
        data = request.get_json(force=True) or {}
        confirm_text = (data.get('confirmText') or '').strip().upper()
        before_date = data.get('beforeDate')  # optional YYYY-MM-DD

        if confirm_text != 'DELETE':
            return jsonify({"success": False, "message": "Confirmation text mismatch. Type DELETE to proceed."}), 400

        # validate date format if provided
        if before_date:
            try:
                datetime.strptime(before_date, '%Y-%m-%d')
            except ValueError:
                return jsonify({"success": False, "message": "Invalid date format. Use YYYY-MM-DD."}), 400

        from services.admin_service import AdminService
        deleted_count = AdminService.delete_old_bookings(before_date)

        # Log activity
        try:
            desc = before_date or 'CURRENT_DATE'
            ActivityService.log_activity(
                action='cleanup',
                entity_type='booking',
                entity_id='bulk',
                entity_name='Old Bookings',
                details=f'Deleted {deleted_count} bookings older than {desc}'
            )
        except Exception:
            pass

        return jsonify({"success": True, "deleted_count": deleted_count})
    except Exception as e:
        logger.error(f"Error cleaning up bookings: {e}")
        return jsonify({"success": False, "message": "Failed to cleanup bookings"}), 500

# Action Routes
@admin_bp.route("/confirm-booking/<booking_id>")
@require_auth
@require_permission('manage_bookings')
def confirm_booking(booking_id):
    """Confirm a booking"""
    return AdminBookingView.confirm_booking(booking_id)

@admin_bp.route("/decline-booking/<booking_id>")
@require_auth
@require_permission('manage_bookings')
def decline_booking(booking_id):
    """Decline a booking"""
    return AdminBookingView.decline_booking(booking_id)

# API Routes
@admin_bp.route("/api/dashboard-stats")
@require_auth
@require_permission('manage_bookings')
def api_dashboard_stats():
    """Get dashboard statistics"""
    return AdminAPIView.get_dashboard_stats()

@admin_bp.route("/api/schedule-data", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_schedule_data():
    """Get schedule data for date range"""
    return AdminAPIView.get_schedule_data()

@admin_bp.route("/api/admin-create-booking", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_admin_create_booking():
    """Create booking from admin panel"""
    return AdminAPIView.create_booking()

@admin_bp.route("/api/update-booking", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_update_booking():
    """Update booking details"""
    return AdminAPIView.update_booking()

@admin_bp.route("/api/admin-booking-action", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_admin_booking_action():
    """Perform action on booking"""
    return AdminAPIView.booking_action()

@admin_bp.route("/api/search-bookings", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_search_bookings():
    """Search bookings by various criteria"""
    return AdminAPIView.search_bookings()

@admin_bp.route("/api/delete-booking", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_delete_booking():
    """Delete booking"""
    return AdminAPIView.delete_booking()

@admin_bp.route("/api/bulk-search", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_bulk_search():
    """Search bookings for bulk operations"""
    return AdminAPIView.bulk_search()

@admin_bp.route("/api/bulk-action", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_bulk_action():
    """Perform bulk action on multiple bookings"""
    return AdminAPIView.bulk_action()

# Reports and Analytics API Routes
@admin_bp.route("/reports")
@require_auth
@require_permission('view_reports')
def reports():
    """Render reports and analytics page"""
    return AdminBookingControlView.render_reports()

@admin_bp.route("/api/reports/dashboard", methods=["GET"])
@require_auth
@require_permission('view_reports')
def api_reports_dashboard():
    """Get dashboard analytics data"""
    return AdminAPIView.get_reports_dashboard()

@admin_bp.route("/api/reports/sports", methods=["GET"])
@require_auth
@require_permission('view_reports')
def api_reports_sports():
    """Get sports performance analytics"""
    return AdminAPIView.get_reports_sports()

@admin_bp.route("/api/reports/revenue", methods=["GET"])
@require_auth
@require_permission('view_reports')
def api_reports_revenue():
    """Get revenue analytics"""
    return AdminAPIView.get_reports_revenue()

@admin_bp.route("/api/reports/customers", methods=["GET"])
@require_auth
@require_permission('view_reports')
def api_reports_customers():
    """Get customer analytics"""
    return AdminAPIView.get_reports_customers()

# Pricing API Routes
@admin_bp.route("/api/pricing", methods=["GET"])
@require_auth
@require_permission('manage_bookings')
def api_get_pricing():
    """Get all court pricing"""
    return AdminAPIView.get_pricing()

@admin_bp.route("/api/pricing/<court_id>", methods=["GET"])
@require_auth
@require_permission('manage_bookings')
def api_get_pricing_by_court(court_id):
    """Get pricing for specific court"""
    return AdminAPIView.get_pricing_by_court(court_id)

@admin_bp.route("/api/pricing", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_create_update_pricing():
    """Create or update court pricing"""
    return AdminAPIView.create_or_update_pricing()

@admin_bp.route("/api/pricing/<court_id>", methods=["DELETE"])
@require_auth
@require_permission('manage_bookings')
def api_delete_pricing(court_id):
    """Delete court pricing"""
    return AdminAPIView.delete_pricing(court_id)

# Promo Code API Routes
@admin_bp.route("/api/promo-codes", methods=["GET"])
@require_auth
@require_permission('manage_bookings')
def api_get_promo_codes():
    """Get all promo codes"""
    try:
        from services.promo_service import PromoService
        promo_codes = PromoService.get_all_promo_codes()
        promo_list = [promo.to_dict() for promo in promo_codes]
        return jsonify({"success": True, "promo_codes": promo_list})
    except Exception as e:
        logger.error(f"Error getting promo codes: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_bp.route("/api/promo-codes/<code>", methods=["GET"])
@require_auth
@require_permission('manage_bookings')
def api_get_promo_code(code):
    """Get specific promo code"""
    try:
        from services.promo_service import PromoService
        promo = PromoService.get_promo_code_by_code(code)
        if promo:
            return jsonify({"success": True, "promo_code": promo.to_dict()})
        else:
            return jsonify({"success": False, "message": "Promo code not found"}), 404
    except Exception as e:
        logger.error(f"Error getting promo code: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_bp.route("/api/promo-codes", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_create_promo_code():
    """Create new promo code"""
    try:
        from services.promo_service import PromoService
        promo_data = request.json
        success, message = PromoService.create_promo_code(promo_data)
        return jsonify({"success": success, "message": message}), (200 if success else 400)
    except Exception as e:
        logger.error(f"Error creating promo code: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_bp.route("/api/promo-codes/<code>", methods=["PUT"])
@require_auth
@require_permission('manage_bookings')
def api_update_promo_code(code):
    """Update existing promo code"""
    try:
        from services.promo_service import PromoService
        promo_data = request.json
        success, message = PromoService.update_promo_code(code, promo_data)
        return jsonify({"success": success, "message": message}), (200 if success else 400)
    except Exception as e:
        logger.error(f"Error updating promo code: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@admin_bp.route("/api/promo-codes/<code>", methods=["DELETE"])
@require_auth
@require_permission('manage_bookings')
def api_delete_promo_code(code):
    """Delete promo code"""
    try:
        from services.promo_service import PromoService
        success, message = PromoService.delete_promo_code(code)
        return jsonify({"success": success, "message": message}), (200 if success else 400)
    except Exception as e:
        logger.error(f"Error deleting promo code: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# Expense API Routes
@admin_bp.route("/api/expenses", methods=["GET"])
@require_auth
@require_permission('manage_bookings')
def api_get_expenses():
    """Get all expenses"""
    return AdminExpenseView.get_all_expenses()

@admin_bp.route("/api/expenses", methods=["POST"])
@require_auth
@require_permission('manage_bookings')
def api_create_expense():
    """Create new expense"""
    return AdminExpenseView.create_expense()

@admin_bp.route("/api/expenses/<int:expense_id>", methods=["PUT"])
@require_auth
@require_permission('manage_bookings')
def api_update_expense(expense_id):
    """Update expense"""
    request.json['id'] = expense_id
    return AdminExpenseView.update_expense()

@admin_bp.route("/api/expenses/<int:expense_id>", methods=["DELETE"])
@require_auth
@require_permission('manage_bookings')
def api_delete_expense(expense_id):
    """Delete expense"""
    request.json = {'id': expense_id}
    return AdminExpenseView.delete_expense()

@admin_bp.route("/api/expenses/statistics")
@require_auth
@require_permission('manage_bookings')
def api_expense_statistics():
    """Get expense statistics"""
    return AdminExpenseView.get_expense_statistics()

@admin_bp.route("/api/expenses/monthly")
@require_auth
@require_permission('manage_bookings')
def api_monthly_expenses():
    """Get monthly expenses"""
    return AdminExpenseView.get_monthly_expenses()

@admin_bp.route("/api/expenses/daily")
@require_auth
@require_permission('manage_bookings')
def api_daily_expenses():
    """Get daily expenses"""
    return AdminExpenseView.get_daily_expenses()

@admin_bp.route("/api/expenses/date-range")
@require_auth
@require_permission('manage_bookings')
def api_date_range_expenses():
    """Get expenses by date range"""
    return AdminExpenseView.get_date_range_expenses()

# Excel Export Routes
@admin_bp.route("/export/bookings")
@require_auth
@require_permission('view_reports')
def export_bookings():
    """Export all bookings to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import make_response
        from io import BytesIO
        from datetime import datetime
        
        # Get all bookings
        bookings = AdminService.get_all_bookings()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Bookings"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Booking ID", "Sport", "Court", "Player Name", "Phone", "Email", 
            "Date", "Start Time", "End Time", "Duration (hrs)", "Player Count",
            "Total Amount", "Status", "Payment Type", "Created At", "Special Requests"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row, booking in enumerate(bookings, 2):
            data = [
                booking.get('id', ''),
                booking.get('sport', ''),
                booking.get('court_name', booking.get('court', '')),
                booking.get('player_name', ''),
                booking.get('player_phone', ''),
                booking.get('player_email', ''),
                booking.get('booking_date', ''),
                booking.get('start_time', ''),
                booking.get('end_time', ''),
                booking.get('duration', 1.0),
                booking.get('player_count', '2'),
                f"PKR {booking.get('total_amount', 0)}",
                booking.get('status', '').title().replace('_', ' '),
                booking.get('payment_type', '').title(),
                booking.get('createdDateTime', ''),
                booking.get('special_requests', '')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=noball_bookings_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Log the export activity
        try:
            from services.activity_service import ActivityService
            ActivityService.log_activity('export', 'bookings', 'all', 'All Bookings', f'Exported {len(bookings)} bookings to Excel')
        except Exception as log_error:
            logger.warning(f"Failed to log export activity: {log_error}")
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting bookings: {e}")
        return jsonify({"error": "Failed to export bookings"}), 500

@admin_bp.route("/export/expenses")
@require_auth
@require_permission('view_expenses')
def export_expenses():
    """Export expenses to Excel with optional filters"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import make_response
        from io import BytesIO
        from datetime import datetime

        # Read optional filters
        area = request.args.get('area')  # a, b, both, all/None
        view = request.args.get('view', 'all')  # all, daily, monthly, range
        date_str = request.args.get('date')  # for daily/monthly
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        # Fetch expenses according to filters
        from services.expense_service import ExpenseService
        expenses = []

        if view == 'daily' and date_str:
            expenses = ExpenseService.get_daily_expenses(date_str, area)
        elif view == 'monthly' and date_str:
            try:
                y, m = map(int, date_str.split('-')[0:2])
                expenses = ExpenseService.get_monthly_expenses(y, m, area)
            except Exception:
                expenses = ExpenseService.get_all_expenses(area_category=area, limit=10000, offset=0)
        elif view == 'range' and start_date and end_date:
            expenses = ExpenseService.get_expenses_by_date_range(start_date, end_date, area)
        else:
            expenses = ExpenseService.get_all_expenses(area_category=area, limit=10000, offset=0)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            "ID", "Title", "Description", "Amount (PKR)", "Category", "Area",
            "Expense Date", "Type", "Frequency", "Created By", "Created At", "Updated At"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for row_idx, exp in enumerate(expenses, 2):
            # Ensure dict access (services returns list of dicts)
            e = exp
            ws.cell(row=row_idx, column=1, value=e.get('id'))
            ws.cell(row=row_idx, column=2, value=e.get('title'))
            ws.cell(row=row_idx, column=3, value=e.get('description'))
            ws.cell(row=row_idx, column=4, value=float(e.get('amount', 0)))
            ws.cell(row=row_idx, column=5, value=e.get('category'))
            ws.cell(row=row_idx, column=6, value=e.get('area_category'))
            ws.cell(row=row_idx, column=7, value=str(e.get('expense_date') or ''))
            ws.cell(row=row_idx, column=8, value=e.get('expense_type'))
            ws.cell(row=row_idx, column=9, value=e.get('recurring_frequency'))
            ws.cell(row=row_idx, column=10, value=e.get('created_by'))
            ws.cell(row=row_idx, column=11, value=str(e.get('created_at') or ''))
            ws.cell(row=row_idx, column=12, value=str(e.get('updated_at') or ''))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=noball_expenses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        # Log export
        try:
            ActivityService.log_activity('export', 'expenses', 'all', 'Expenses', f'Exported {len(expenses)} expenses')
        except Exception:
            pass

        return response
    except Exception as e:
        logger.error(f"Error exporting expenses: {e}")
        return jsonify({"error": "Failed to export expenses"}), 500
